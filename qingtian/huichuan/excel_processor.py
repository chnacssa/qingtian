"""汇川 Excel 处理器 — Sheet 独立编译 + 图片/图表提取

Phase 3 模块。将多 Sheet XLSX 文件逐 Sheet 解析为 Markdown 表格，
提取嵌入图片，支持独立编译（每 Sheet 一条 knowledge_entry）
和合并编译（全部 Sheet 拍扁为一个 Markdown）。

与 _extract_xlsx() (ingest.py) 的关系：
  - _extract_xlsx 简单提取所有单元格文本（已有，不改动）
  - 本模块提供结构化 Markdown 表格 + 图片提取（新增）
  - 通过 config.get_excel_sheet_independent() 控制启用/回退
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
from datetime import date

import openpyxl

from . import config as kcfg
from .image_extractor import ImageRecord, detect_format_from_bytes

logger = logging.getLogger("huichuan.excel_processor")

# ── 边界常量 ────────────────────────────────────────────

MAX_SHEETS = 20         # 最多处理的 Sheet 数
MAX_IMAGES_PER_SHEET = 10


class SheetResult:
    """单个 Sheet 的解析结果。

    Attributes:
        sheet_name: Sheet 名
        row_count: 数据行数（含标题行）
        col_count: 列数
        markdown: Markdown 表格文本
        images: 提取到的嵌入图片列表
    """
    def __init__(self, sheet_name: str, row_count: int, col_count: int,
                 markdown: str, images: list[ImageRecord]):
        self.sheet_name = sheet_name
        self.row_count = row_count
        self.col_count = col_count
        self.markdown = markdown
        self.images = images


async def process_xlsx(data: bytes, storage_base: str,
                       file_id: str,
                       extract_images: bool = True) -> list[SheetResult]:
    """XLSX 完整解析：逐 Sheet → Markdown + 嵌入图片提取。

    在线程池内执行全部 I/O 密集型操作。

    Args:
        data: XLSX 文件原始字节
        storage_base: Layer 1 根目录（用于图片存储）
        file_id: 文件注册表 ID（用于图片子目录）

    Returns:
        list[SheetResult] — 解析结果列表（密码保护/损坏时返回空列表）
    """
    def _process() -> list[SheetResult]:
        try:
            wb = openpyxl.load_workbook(data, data_only=True)
        except Exception as e:
            _log_open_error(e)
            return []

        results: list[SheetResult] = []
        for si, sheet_name in enumerate(wb.sheetnames):
            if si >= MAX_SHEETS:
                logger.warning("XLSX '%s' has more than %d sheets, truncating",
                              sheet_name, MAX_SHEETS)
                break
            ws = wb[sheet_name]

            # ── Markdown 表格 ────────────────────────────
            rows_list = list(ws.iter_rows(values_only=True))
            if not rows_list:
                results.append(SheetResult(sheet_name, 0, 0, "", []))
                continue

            def _md_cell(v) -> str:
                """转义 Markdown 表格单元格中的特殊字符。"""
                s = str(v) if v is not None else ""
                # 管道符 → HTML 实体，避免破坏表格列结构
                s = s.replace("|", "&#124;")
                # 换行 → 空格（Markdown 表格不支持单元格内换行）
                s = s.replace("\n", " ").replace("\r", " ")
                return s

            headers = [_md_cell(c) for c in rows_list[0]]
            md_lines = [
                "| " + " | ".join(headers) + " |",
                "| " + " | ".join("---" for _ in headers) + " |",
            ]
            for row in rows_list[1:]:
                cells = [_md_cell(c) for c in row]
                md_lines.append("| " + " | ".join(cells) + " |")

            # ── 嵌入图片提取 ────────────────────────────
            images: list[ImageRecord] = []
            if extract_images and hasattr(ws, "_images"):
                for ii, img in enumerate(ws._images or []):
                    if ii >= MAX_IMAGES_PER_SHEET:
                        break
                    raw = _extract_image_data(img)
                    if raw is None or len(raw) > kcfg.get_max_image_size_mb() * 1024 * 1024:
                        continue
                    xlsx_fmt = detect_format_from_bytes(raw)
                    record = ImageRecord(
                        raw, xlsx_fmt,
                        image_index=len(images),
                        source_sheet=sheet_name,
                    )
                    record.storage_path = _save_xlsx_image(
                        raw, xlsx_fmt, storage_base, file_id, si, len(images),
                    )
                    images.append(record)

            results.append(SheetResult(
                sheet_name, len(rows_list), len(headers),
                "\n".join(md_lines), images,
            ))

        wb.close()
        return results

    return await asyncio.to_thread(_process)


async def xlsx_to_entries(data: bytes, source: str, filename: str,
                          storage_path: str, conn, schema: str,
                          storage_base: str = "",
                          sheets: list[SheetResult] | None = None) -> list[dict]:
    """XLSX → 按配置模式编译为 knowledge_entry。

    根据 config.get_excel_sheet_independent() 决定模式：
      - True（默认）: 每 Sheet 独立入库，保留源 Sheet 上下文
      - False: 所有 Sheet 合并为一个 Markdown 文档

    Args:
        data: XLSX 文件原始字节
        source: 来源标识
        filename: 原始文件名
        storage_path: Layer 1 存储路径
        conn: asyncpg connection
        schema: 数据库 schema
        storage_base: Layer 1 根目录（用于图片存储）
        sheets: 可复用调用方已完成的解析结果（P2 R11，避免同文件二次全量解析）

    Returns:
        list[dict] — ingest_text 返回值列表
    """
    from huichuan.ingest import ingest_text

    sheet_independent = kcfg.get_excel_sheet_independent()
    results: list[dict] = []

    # P2 (R11): 复用调用方传入的解析结果（ingest_file 已用 extract_images=True
    # 全量解析过一次）；未提供时自行解析（向后兼容独立调用）。
    if sheets is None:
        sheets = await process_xlsx(data, storage_base, "", extract_images=False)

    if not sheet_independent:
        # 合并模式：所有 Sheet 平铺为一个文档
        parts = []
        for s in sheets:
            if s.markdown:
                parts.append(f"## Sheet: {s.sheet_name}\n{s.markdown}")
        if parts:
            combined = "\n\n".join(parts)
            result = await ingest_text(conn, combined, source=source,
                                       original_filename=filename,
                                       storage_path=storage_path, schema=schema)
            results.append(result)
    else:
        # 独立模式：每 Sheet 单独入库
        for sheet in sheets:
            if not sheet.markdown:
                continue
            sheet_source = f"{source}:{sheet.sheet_name}"
            result = await ingest_text(
                conn, sheet.markdown,
                source=sheet_source,
                original_filename=f"{filename}[{sheet.sheet_name}]",
                storage_path=storage_path,
                schema=schema,
            )
            if result.get("knowledge_ids"):
                for kid in result["knowledge_ids"]:
                    await conn.execute(
                        f"UPDATE {schema}.knowledge_entries "
                        f"SET metadata = metadata || $1::jsonb "
                        f"WHERE knowledge_id = $2",
                        json.dumps({
                            "sheet_name": sheet.sheet_name,
                            "row_count": sheet.row_count,
                            "col_count": sheet.col_count,
                        }, ensure_ascii=False),
                        kid,
                    )
            results.append(result)

    return results


# ── 内部工具函数 ────────────────────────────────────────


def _extract_image_data(img) -> bytes | None:
    """从 openpyxl 图片对象提取字节数据。

    兼容不同 openpyxl 版本的 _data() / ref / blob 接口。
    """
    raw = None
    if hasattr(img, "_data"):
        try:
            raw = img._data()
        except Exception:
            pass
    elif hasattr(img, "ref"):
        raw = img.ref
    elif hasattr(img, "blob"):
        raw = img.blob

    if isinstance(raw, bytes):
        return raw
    if hasattr(raw, "read"):
        try:
            return raw.read()
        except Exception:
            return None
    return None


def _save_xlsx_image(data: bytes, fmt: str, storage_base: str,
                     file_id: str, sheet_idx: int, img_idx: int) -> str:
    """保存 XLSX 嵌入图片到 Layer 1。

    路径: storage/{year}/{month}/images/{file_id}/s{sheet_idx}_{img_idx}.{fmt}
    """
    today = date.today()
    img_dir = os.path.join(storage_base, str(today.year), f"{today.month:02d}",
                          "images", file_id)
    os.makedirs(img_dir, exist_ok=True)
    path = os.path.join(img_dir, f"s{sheet_idx}_{img_idx}.{fmt}")
    with open(path, "wb") as f:
        f.write(data)
    return path


def _log_open_error(e: Exception) -> None:
    """统一处理 XLSX 打开失败日志。"""
    err = str(e).lower()
    if any(kw in err for kw in ("password", "encrypted", "mso")):
        logger.warning("XLSX is password-protected or encrypted: %s", e)
    else:
        logger.warning("XLSX open failed: %s", e)
