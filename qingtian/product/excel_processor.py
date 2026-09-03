"""
产品目录/价目表 XLSX 批量导入

依赖：openpyxl

目录导入格式（必填列用 * 标记）：
  *类别, *产品名称, 型号, 电压等级, 容量, 执行标准, 单位, 单价(元)

价目表导入格式：
  *产品名称, 型号, *单价(元), 币种, 数量折扣(JSON)
"""

import io
import json
import logging

logger = logging.getLogger("product.excel_processor")


# ── 产品目录导入 ──────────────────────────────────


# 列索引映射
COL_CATEGORY = 0
COL_NAME = 1
COL_MODEL = 2
COL_VOLTAGE = 3
COL_POWER = 4
COL_STANDARDS = 5
COL_UNIT = 6


def parse_catalog_xlsx(file_bytes: bytes, enterprise_id: str, created_by: str = "") -> dict:
    """解析产品目录 XLSX，返回结构化数据列表。

    返回:
        {
            "total": int,
            "products": [dict, ...],
            "errors": [str, ...],
        }
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return {"total": 0, "products": [], "errors": ["工作簿中没有工作表"]}

    products = []
    errors = []
    row_num = 0

    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        row_num += 1
        if not row or not any(v is not None for v in row):
            continue  # 跳过空行

        try:
            category = str(row[COL_CATEGORY] or "").strip()
            name = str(row[COL_NAME] or "").strip()

            if not category:
                errors.append(f"第 {row_num + 1} 行：类别为空，跳过")
                continue
            if not name:
                errors.append(f"第 {row_num + 1} 行：产品名称为空，跳过")
                continue

            # 解析执行标准（逗号分隔或分号分隔）
            standards_str = str(row[COL_STANDARDS] or "").strip()
            standards = [s.strip() for s in standards_str.replace("；", ";").split(";") if s.strip()] if standards_str else []

            product = {
                "enterprise_id": enterprise_id,
                "category": category,
                "name": name,
                "model": str(row[COL_MODEL] or "").strip(),
                "voltage_level": str(row[COL_VOLTAGE] or "").strip(),
                "power_rating": str(row[COL_POWER] or "").strip(),
                "standards": standards,
                "unit": str(row[COL_UNIT] or "台").strip(),
                "created_by": created_by,
            }
            products.append(product)
        except Exception as e:
            errors.append(f"第 {row_num + 1} 行解析失败: {e}")

    wb.close()
    logger.info("Catalog XLSX parsed: %d products, %d errors", len(products), len(errors))
    return {"total": len(products), "products": products, "errors": errors}


# ── 价目表导入 ────────────────────────────────────


COL_PNAME = 0
COL_PMODEL = 1
COL_UNIT_PRICE = 2
COL_CURRENCY = 3
COL_QTY_DISCOUNT = 4


def parse_price_list_xlsx(file_bytes: bytes, price_list_id: str) -> dict:
    """解析价目表 XLSX，返回明细行数据列表。

    返回:
        {
            "total": int,
            "items": [dict, ...],
            "errors": [str, ...],
        }
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return {"total": 0, "items": [], "errors": ["工作簿中没有工作表"]}

    items = []
    errors = []
    row_num = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if not row or not any(v is not None for v in row):
            continue

        try:
            product_name = str(row[COL_PNAME] or "").strip()
            unit_price = row[COL_UNIT_PRICE]

            if not product_name:
                errors.append(f"第 {row_num + 1} 行：产品名称为空，跳过")
                continue

            try:
                price = float(unit_price) if unit_price is not None else 0
                if price <= 0:
                    errors.append(f"第 {row_num + 1} 行：单价必须大于 0，跳过")
                    continue
            except (ValueError, TypeError):
                errors.append(f"第 {row_num + 1} 行：单价格式无效 '{unit_price}'，跳过")
                continue

            # 解析数量折扣（JSON 字符串）
            qty_discount = {}
            qty_str = str(row[COL_QTY_DISCOUNT] or "").strip()
            if qty_str:
                try:
                    qty_discount = json.loads(qty_str)
                    if not isinstance(qty_discount, dict):
                        qty_discount = {}
                except json.JSONDecodeError:
                    errors.append(f"第 {row_num + 1} 行：数量折扣 JSON 格式无效，跳过折扣")
                    qty_discount = {}

            item = {
                "price_list_id": price_list_id,
                "product_spec": {
                    "name": product_name,
                    "model": str(row[COL_PMODEL] or "").strip() if len(row) > COL_PMODEL else "",
                },
                "unit_price": price,
                "currency": str(row[COL_CURRENCY] or "CNY").strip() if len(row) > COL_CURRENCY else "CNY",
                "quantity_discount": qty_discount,
                "sort_order": row_num,
            }
            items.append(item)
        except Exception as e:
            errors.append(f"第 {row_num + 1} 行解析失败: {e}")

    wb.close()
    logger.info("PriceList XLSX parsed: %d items, %d errors", len(items), len(errors))
    return {"total": len(items), "items": items, "errors": errors}
